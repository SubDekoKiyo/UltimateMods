import os
from openpyxl import load_workbook

print("Please Choose Language Number")
print("")
print("1: English")
print("2: Latam")
print("3: Brazilian")
print("4: Portuguese")
print("5: Korean")
print("6: Russian")
print("7: Dutch")
print("8: Filipino")
print("9: French")
print("10: German	")
print("11: Italian")
print("12: Japanese")
print("13: Spanish")
print("14: SChinese")
print("15: TChinese")
print("16: Irish")
print("")
print("Please Choose 1~16")
LanguageNum = int(input())

LanguageText = "English"

if (LanguageNum == 1):
  LanguageText = "English"
elif (LanguageNum == 2):
  LanguageText = "Latam"
elif (LanguageNum == 3):
  LanguageText = "Brazilian"
elif (LanguageNum == 4):
  LanguageText = "Portuguese"
elif (LanguageNum == 5):
  LanguageText = "Korean"
elif (LanguageNum == 6):
  LanguageText = "Russian"
elif (LanguageNum == 7):
  LanguageText = "Dutch"
elif (LanguageNum == 8):
  LanguageText = "Filipino"
elif (LanguageNum == 9):
  LanguageText = "French"
elif (LanguageNum == 10):
  LanguageText = "German"
elif (LanguageNum == 11):
  LanguageText = "Italian"
elif (LanguageNum == 12):
  LanguageText = "Japanese"
elif (LanguageNum == 13):
  LanguageText = "Spanish"
elif (LanguageNum == 14):
  LanguageText = "SChinese"
elif (LanguageNum == 15):
  LanguageText = "TChinese"
elif (LanguageNum == 16):
  LanguageText = "Irish"

WORKING_DIR = os.path.dirname(os.path.realpath(__file__))
IN_FILE = os.path.join(WORKING_DIR, "Translate.xlsx")
OUT_FILE = os.path.join(WORKING_DIR, f"Language{LanguageText}.dat")

f = open(OUT_FILE, "w", newline="\n")
f.write("")
f.close()

def GenerateLanguageFile(InFiles):
  for FileName in InFiles:
    if not os.path.isfile(FileName):
      continue

    WorkBook = load_workbook(FileName, read_only = True)

    for StringData in WorkBook:
      Rows = StringData.iter_rows(min_col = 1, min_row = 2, max_col = LanguageNum+1, max_row = None)
      Headers = []
      for Header in StringData[1]:
        if Header.value:
          Headers.append(Header.value)

      for Row in Rows:
        Name = Row[0].value

        if not Name:
          continue

        Data = ""

        for string in Row[LanguageNum-1:]:
          if string.value:
            Data = string.value.replace("\\r", "\r").replace("_x000D_", "").replace("\\n", "\n").encode('unicode-escape').decode('ascii')

        WriteData = [f'{str(Name)}',':',f'{str(Data)}','\n']

        f = open(OUT_FILE, "a", newline="\n", encoding="utf-8_sig")
        f.writelines(WriteData)
        f.close()

if __name__ == "__main__":
  InFiles = [
    os.path.join(IN_FILE)
  ]

  GenerateLanguageFile(InFiles)